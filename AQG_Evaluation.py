import os
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import language_tool_python
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer
from sentence_transformers import SentenceTransformer
from langchain_groq import ChatGroq
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.chains.combine_documents import create_stuff_documents_chain
from langchain_core.prompts import ChatPromptTemplate
from langchain.chains import create_retrieval_chain
from langchain_community.vectorstores import FAISS
from langchain_community.document_loaders import WebBaseLoader
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from textstat.textstat import textstat
from openpyxl import load_workbook
import re
import bs4
from dotenv import load_dotenv
from scipy.stats import pearsonr
from statsmodels.stats.inter_rater import fleiss_kappa
import krippendorff
import rbo
from scipy.stats import rankdata

# Load environment variables
load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")
os.environ["GOOGLE_API_KEY"] = os.getenv("GOOGLE_API_KEY")
os.environ['USER_AGENT'] = 'myagent'

# Configure Streamlit
st.set_page_config(layout='wide', page_title="Question Generation and Evaluation")

# Model configurations
MODELS = [
    {"name": "Gemma", "model_name": "gemma2-9b-it"},
    {"name": "Llama", "model_name": "llama3-8b-8192"},
    {"name": "Mistral", "model_name": "mixtral-8x7b-32768"}
]

embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
file_path = "Output/generated_questions.xlsx"

# Prompt template
prompt_template = ChatPromptTemplate.from_template(
    """
    Answer the questions based on the provided context only.
    <context>{context}</context>
    Questions:{input}
    """
)
prompt1 = """
    Generate questions for all six levels of Bloom's Taxonomy. 
    - Include as many questions as possible for each level.
    - Do not create multiple-choice questions.
    - Use the levels: Remembering, Understanding, Applying, Analyzing, Evaluating, Creating.
"""

# Define Bloom's Taxonomy Levels and Associated Keywords
BLOOMS_KEYWORDS = {
    "Remembering": ["define", "find", "how", "identify", "label", "list", "match", "name", "omit", "recall", 
                    "relate", "select", "show", "spell", "state","tell", "what", "when", "where", "which", "who", "why"],
    "Understanding": ["classify", "compare", "contrast", "demonstrate", "describe", "explain", "extend", 
                      "illustrate", "infer", "interpret", "outline", "relate", "rephrase", "show", "summarize", "translate"],
    "Applying": ["apply", "build", "choose", "construct", "demonstrate", "develop", "experiment with", "identify", 
                 "illustrate", "implement", "interview", "make use of", "model", "organize", "plan", "select", "show", "solve", "use", "utilize"],
    "Analyzing": ["analyze", "assume", "break down", "categorize", "classify", "compare", "conclusion", "contrast", 
                  "differentiate", "discover", "dissect", "distinguish", "divide", "examine", "function", "inference", 
                  "inspect", "list", "motive", "relationships", "simplify", "survey", "take part in", "test for", "theme"],
    "Evaluating": ["agree", "appraise", "argue", "assess", "award", "choose", "compare", "conclude", "criteria", "criticize", 
                   "decide", "deduct", "defend", "determine", "disprove","estimate", "evaluate", "explain", "importance", 
                   "influence", "interpret", "judge", "justify", "mark", "measure", "opinion", "perceive", "prioritize", 
                   "prove", "rate", "recommend", "review", "rule on", "select", "support", "value"],
    "Creating": ["adapt", "build", "change", "choose", "combine", "compile", "compose", "construct", "create", "delete", 
                 "design", "develop", "discuss", "elaborate", "estimate", "formulate", "happen", "imagine", "improve", 
                 "invent", "make up", "maximize", "minimize", "modify", "original", "originate", "plan", "predict","propose",
                  "solution", "solve", "suppose", "test", "theory"]
}

# Global serial number for questions
sr_no = 1

tool = language_tool_python.LanguageTool('en-US')

# Function: Vector Embedding
def vector_embedding(url):
    """Loads and splits context into vector embeddings."""
    if "vectors" not in st.session_state or st.session_state.get("url") != url:
        try:
            st.session_state.embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
            loader = WebBaseLoader(
                web_paths=[url],
                bs_kwargs={"parse_only": bs4.SoupStrainer(class_="chapter-content-module")},
                bs_get_text_kwargs={"separator": " | ", "strip": True},
            )
            st.session_state.docs = loader.load()
            text_splitter = RecursiveCharacterTextSplitter(chunk_size=5000, chunk_overlap=200)
            st.session_state.final_documents = text_splitter.split_documents(st.session_state.docs)
            st.session_state.vectors = FAISS.from_documents(st.session_state.final_documents, st.session_state.embeddings)
            st.session_state.url = url
        except Exception as e:
            st.error(f"Error processing the URL: {e}")
            raise

def check_answerability(question, context, threshold=0.6):  # Lowered the threshold 
    """
    Checks if a given question can be answered based on the context using semantic similarity.
    
    Args:
        question (str): The question to evaluate.
        context (str): The reference context.
        threshold (float): Similarity score threshold (default: 0.5).

    Returns:
        int: 1 if the question is answerable, otherwise 0.
    """
    # Ensure the context is split into smaller chunks for better matching
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=300, chunk_overlap=50)
    context_chunks = text_splitter.split_text(context)

    max_similarity = 0  # Track the highest similarity found
    
    # Compare question with each small chunk of context
    question_embedding = embedding_model.encode([question])
    
    for chunk in context_chunks:
        chunk_embedding = embedding_model.encode([chunk])
        similarity_score = cosine_similarity(question_embedding, chunk_embedding)[0][0]
        max_similarity = max(max_similarity, similarity_score)  # Keep the highest similarity
    
    return 1 if max_similarity >= threshold else 0  # Answerable if max similarity is high



# Function: Create DataFrame
def create_dataframe(response_text, url, model_name):
    """Parses the model response into a structured DataFrame."""
    global sr_no
    data = []
    match = re.search(r'books\/(.*?)\/', url)
    if match:
        subject = match.group(1).capitalize()
    else:
        subject = "Subject not found."	
    print(subject)
    match = re.search(r'pages/([\d-]+)-(.+)', url)
    if match:
        chapter_number = match.group(1).replace("-", ".")  # Replace '-' with '.' for chapter number
        chapter_title = match.group(2).replace("-", " ").title()  # Replace '-' with space and capitalize
        chapter_name = f"{chapter_number} {chapter_title}"
    else:
        chapter_name = "Chapter name not found."
    lines = response_text.strip().split("\n")
    current_level = None
    context_text = " ".join([doc.page_content for doc in st.session_state.final_documents])
    for line in lines:
        if line.startswith("**") and line.endswith("**") or line.endswith(":"):
            current_level = re.sub(r'[^a-zA-Z0-9]', '', line)
        elif current_level and line.strip():
            question_match = re.match(r"(.*?)(\s\(Confidence: ([0-9\.]+)\))?$", line.strip("*"))
            question_text = question_match.group(1) if question_match else line.strip("*")
            
            # Readability Calculation
            readability = 1 if (textstat.flesch_reading_ease(question_text)) > 50 else 0

             # Grammar Check using LanguageTool
            grammar_errors = len(tool.check(question_text))
            grammar_correct = 1 if grammar_errors == 0 else 0  # 1 if correct, 0 if incorrect

            # Context Retention Check
            context_retension = check_context_retention(question_text, context_text)

            # Evaluate Bloom’s Level
            blooms_level_correct = evaluate_blooms_level(question_text, current_level)

            # Answerability Check
            answerable = check_answerability(question_text, context_text)

            data.append({
                "Sr.No.": sr_no,
                "Subject" : subject,
                "Chapter Name": chapter_name,
                "Model Name": model_name,                
                "Bloom's Level": current_level,
                "Question": question_text,
                "Word Count": len(question_text.split()),
                "Clarity": readability,
                "Grammar": grammar_correct,
                "Relevance ": context_retension,
                "Bloom's Level Fit": blooms_level_correct,
                "Answerable": answerable
            })
            sr_no += 1
    
    df = pd.DataFrame(data)
    save_to_excel(df, file_path)
    return df

# Function: Save to Excel
def save_to_excel(df, file_path):
    """Save data to an Excel file. Append if the file exists."""
    if not os.path.exists("Output"):
        os.makedirs("Output")
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            sheet = workbook["Sheet1"]
            start_row = sheet.max_row
            df.to_excel(writer, index=False, header=False if start_row > 0 else True, startrow=start_row, sheet_name='Sheet1')
    else:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

combined_df = []
# Function: Generate Questions
def generate_questions(url):
    """Generates questions for all models and saves them to an Excel file."""
    vector_embedding(url)
    global combined_df
    for model in MODELS:
        try:
            st.write(f"Generating questions for {model['name']}...")
            llm = ChatGroq(groq_api_key=groq_api_key, model_name=model["model_name"])
            retriever = st.session_state.vectors.as_retriever()
            retrieval_chain = create_retrieval_chain(retriever, create_stuff_documents_chain(llm, prompt_template))
            response = retrieval_chain.invoke({'input': prompt1})
            df = create_dataframe(response["answer"], url, model["name"])
            combined_df.append(df)
            st.dataframe(df)
        except Exception as e:
            st.error(f"Error generating questions for {model['name']}: {e}")
    return pd.concat(combined_df)

def check_context_retention(question, context, threshold=0.7):
    """
    Check if a question belongs to the given context using cosine similarity.

    Args:
        question (str): The question to evaluate.
        context (str): The original context.
        threshold (float): Similarity score threshold for determining context match.

    Returns:
        int: 1 if the question belongs to the context, else 0.
    """
    question_embedding = embedding_model.encode([question])
    
    # Split context into sentences to improve local relevance
    context_sentences = context.split(". ")
    context_embeddings = embedding_model.encode(context_sentences)

    # Compute cosine similarity between question and each sentence in the context
    similarity = cosine_similarity(question_embedding, context_embeddings)[0]

    # Check if any similarity score is above the threshold
    if np.max(similarity) >= threshold:
        return 1  # Question is contextually relevant
    else:
        return 0  # Question does not belong to the context

def evaluate_blooms_level(question, assigned_level):
    """
    Evaluates if the assigned Bloom's Taxonomy level is correct by checking the presence
    of corresponding keywords in the question.

    Args:
    - question (str): The question text.
    - assigned_level (str): The Bloom's Taxonomy level assigned to the question.

    Returns:
    - (int): 1 if the assigned level is correct, 0 otherwise.
    """
    # Convert assigned Bloom’s level to match dictionary keys
    assigned_level = assigned_level.strip().capitalize()

    # Get the expected keywords for the assigned level
    expected_keywords = BLOOMS_KEYWORDS.get(assigned_level, [])

    # Check if at least one expected keyword is in the question
    for keyword in expected_keywords:
        if keyword in question.lower():
            return 1  # Question correctly matches Bloom’s level

    return 0  # Question does not align with expected Bloom’s level

# Utility Functions for Analysis
def compute_semantic_similarity(questions):
    embeddings = embedding_model.encode(questions)
    pairwise_similarity = cosine_similarity(embeddings)
    avg_similarity = (pairwise_similarity.sum() - len(questions)) / (len(questions) * (len(questions) - 1)) if len(questions) > 1 else 1
    return avg_similarity

def compute_jaccard_diversity(questions):
    """Computes Jaccard diversity for the given list of questions."""
    unique_terms = [set(q.split()) for q in questions]
    pairwise_jaccard = []
    for i, a in enumerate(unique_terms):
        for b in unique_terms[i + 1:]:
            union_len = len(a | b)
            if union_len > 0:  # Avoid division by zero
                pairwise_jaccard.append(len(a & b) / union_len)
    if pairwise_jaccard:  # If there are valid pairs
        return 1 - (sum(pairwise_jaccard) / len(pairwise_jaccard))
    return 1  # Maximum diversity if no pairs exist

def compute_cosine_diversity(questions):
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(questions)
    pairwise_cosine = cosine_similarity(tfidf_matrix)
    avg_similarity = (pairwise_cosine.sum() - len(questions)) / (len(questions) * (len(questions) - 1)) if len(questions) > 1 else 1
    return 1 - avg_similarity

# Function to compute context utilization
def compute_context_utilization(df, context):
    context_terms = set(context.split())
    question_terms = [set(question.split()) for question in df["Question"]]
    utilization = [len(context_terms & terms) / len(context_terms) for terms in question_terms]
    return np.mean(utilization)

# Function to Evaluate Grammar Quality
def evaluate_grammar(questions):
    """Evaluate grammatical correctness of questions."""
    #errors = [len(tool.check(question)) for question in questions]
    #return np.mean(errors)  # Average grammatical errors per question
    for question in questions:
        errors = tool.check(question)
        if errors:  # If there are any grammar issues
            print(question)
    
    return np.mean([len(tool.check(question)) for question in questions])  # Average grammatical errors per question


def plot_readability_distribution(questions):
    scores = [textstat.flesch_reading_ease(q) for q in questions]
    fig, ax = plt.subplots(figsize=(5, 3), dpi=100)
    ax.hist(scores, bins=10, alpha=0.7)
    ax.set_title("Readability Distribution", fontsize=10)
    ax.set_xlabel("Flesch Reading Ease Score", fontsize=8)
    ax.set_ylabel("Frequency", fontsize=8)
    st.pyplot(fig)

def plot_blooms_distribution(df):
    fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
    bloom_distribution = df.groupby(["Model Name", "Bloom's Level"]).size().unstack(fill_value=0)
    bloom_distribution.plot(kind="bar", stacked=True, ax=ax)
    ax.set_title("Bloom's Level Distribution", fontsize=10)
    ax.set_ylabel("Number of Questions", fontsize=8)
    ax.set_xlabel("Model Name", fontsize=8)
    ax.tick_params(axis='x', labelsize=8, rotation=0)  # Adjust x-axis tick labels
    ax.legend().set_visible(False)
    # Adjust legend placement
    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels, loc='lower center', bbox_to_anchor=(0.5, -0.15), 
               ncol=3, fontsize=8, frameon=False)
    st.pyplot(fig)

def plot_question_similarity_heatmap(combined_df):
    model_questions = combined_df.groupby("Model Name")["Question"].apply(list)
    model_names = model_questions.index
    num_models = len(model_names)
    similarity_matrix = np.zeros((num_models, num_models))
    for i in range(num_models):
        for j in range(num_models):
            if i <= j:
                questions_i = " ".join(model_questions[model_names[i]])
                questions_j = " ".join(model_questions[model_names[j]])
                embeddings_i = embedding_model.encode([questions_i])
                embeddings_j = embedding_model.encode([questions_j])
                similarity = cosine_similarity(embeddings_i, embeddings_j)[0, 0]
                similarity_matrix[i, j] = similarity
                similarity_matrix[j, i] = similarity
    similarity_df = pd.DataFrame(similarity_matrix, index=model_names, columns=model_names)
    fig, ax = plt.subplots(figsize=(6, 5), dpi=100)
    sns.heatmap(similarity_df, annot=True, fmt=".2f", cmap="coolwarm", cbar=True, ax=ax)
    ax.set_title("Question Similarity Heatmap", fontsize=12)
    st.pyplot(fig)

# Visualization: Radar Chart
def plot_radar_chart(metrics_df, metrics_to_plot):
    """
    Plots a radar chart comparing models across multiple metrics.
    Args:
        metrics_df (pd.DataFrame): DataFrame containing metrics for all models.
        metrics_to_plot (list): List of metric names to include in the radar chart.
    """
    normalized_df = metrics_df.copy()
    for metric in metrics_to_plot:
        normalized_df[metric] = normalized_df[metric] / normalized_df[metric].max()

    labels = metrics_to_plot
    num_vars = len(labels)

    # Compute angles for radar chart
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    angles += angles[:1]

    # Create the radar chart
    fig, ax = plt.subplots(figsize=(4, 4), subplot_kw=dict(polar=True), dpi=100)
    for i, row in normalized_df.iterrows():
        values = row[metrics_to_plot].tolist()
        values += values[:1]
        ax.plot(angles, values, label=row["model_name"])
        ax.fill(angles, values, alpha=0.25)

    ax.set_yticks([])
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=8, wrap=True)
    ax.set_title("Radar Chart: Model Comparison", va="top", fontsize=10, pad=20)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), ncol=3, fontsize=8, frameon=False)
    st.pyplot(fig)
    st.write("\n\n")

# Function: Evaluate Models
def evaluate_models(combined_df):
    """Evaluates models and generates metrics and visualizations."""
    st.write("### Evaluating Models...")
    metrics_list = []
    
    for model in automated_df["Model Name"].unique():
        model_df = automated_df[automated_df["Model Name"] == model]
        questions = model_df["Question"].tolist()

        metrics = {
            "Model Name": model,
            "Semantic Similarity": compute_semantic_similarity(questions),
            "Jaccard Diversity": compute_jaccard_diversity(questions),
            "Cosine Diversity": compute_cosine_diversity(questions),
            "Context Utilization": model_df["Relevance"].mean(),  # Already computed in dataset
            "Grammar Errors": 1 - model_df["Grammar"].mean(),  # Convert correctness to errors
            "Readability": model_df["Clarity"].mean(),  # Clarity represents readability
            "Answerability": model_df["Answerable"].mean(),  # Proportion of answerable questions
            "Bloom’s Level Fit": model_df["Bloom's Level Fit"].mean(),  # Accuracy of Bloom’s level classification
        }

        metrics_list.append(metrics)

    metrics_df = pd.DataFrame(metrics_list)

# Display results
    st.write("### 📊 Model-wise Automated Evaluation Metrics")
    st.dataframe(metrics_df)

    # **Identify Best Model**
    best_model = metrics_df.loc[metrics_df["Semantic Similarity"].idxmax(), "Model Name"]
    best_similarity = metrics_df["Semantic Similarity"].max()

    st.write(f"### 🏆 Best Model Based on Automated Evaluation: **{best_model}**")
    st.write(f"Highest Semantic Similarity Score: **{best_similarity:.3f}**")

    # **Provide Reasoning for Best Model**
    st.write("#### 📌 Why is this the Best Model?")
    st.write(f"- {best_model} achieved the **highest semantic similarity**, indicating better question relevance.")
    st.write("- High **context utilization** suggests the model effectively used provided content.")
    st.write("- Lower **grammar errors** indicate higher linguistic quality.")

    # **Visualization - Grouped Bar Chart**
    fig, ax = plt.subplots(figsize=(10, 5))
    metrics_df.plot(x="Model Name", kind="bar", ax=ax, colormap="viridis")

    ax.set_title("Model-wise Automated Evaluation Metrics")
    ax.set_ylabel("Score")
    ax.set_xticklabels(metrics_df["Model Name"], rotation=0)
    ax.legend(loc="upper left")

    st.pyplot(fig)

    # Radar Chart Visualization
    st.write("##### Radar Chart: Provides a comparative overview of all key metrics for the models in a single visualization.")
    plot_radar_chart(combined_df, [
         "Semantic Similarity", "Jaccard Diversity", 
        "Cosine Diversity", "Context Utilization", "Grammar Errors", 
        "Readability", "Answerability", "Bloom’s Level Fit"
    ])
    plot_blooms_distribution(combined_df)
    plot_readability_distribution(combined_df["Question"].tolist())
    plot_question_similarity_heatmap(combined_df)


def compute_fleiss_kappa_per_model(sme_df):
    """
    Compute Fleiss' Kappa for each evaluation criterion per model, export results to Excel, 
    and determine the best-performing model.

    Args:
        sme_df (pd.DataFrame): DataFrame containing SME evaluations 
                               for Clarity, Grammar, Relevance, 
                               Bloom’s Level Fit, and Answerability.

    Returns:
        pd.DataFrame: Fleiss' Kappa scores per model for each evaluation criterion.
    """

    # Define evaluation criteria and corresponding SME columns
    evaluation_criteria = {
        "Clarity": ["Clarity_SME1", "Clarity_SME2", "Clarity_SME3"],
        "Grammar": ["Grammar_SME1", "Grammar_SME2", "Grammar_SME3"],
        "Relevance": ["Relevance_SME1", "Relevance_SME2", "Relevance_SME3"],
        "Bloom’s Level Fit": ["Bloom's Level Fit_SME1", "Bloom's Level Fit_SME2", "Bloom's Level Fit_SME3"],
        "Answerability": ["Answerable_SME1", "Answerable_SME2", "Answerable_SME3"]
    }

    # Store results
    kappa_scores = []

    # Compute Fleiss' Kappa for each model
    for model in sme_df["Model Name"].unique():
        model_data = sme_df[sme_df["Model Name"] == model]
        model_kappa_scores = {"Model Name": model}

        for criterion, sme_columns in evaluation_criteria.items():
            # Extract SME ratings for the current model and criterion
            ratings = model_data[sme_columns].values

            # Ensure there are at least two distinct values (0 and 1) in the dataset
            unique_values = np.unique(ratings)
            if len(unique_values) == 1:  # If all ratings are the same, assume perfect agreement
                kappa_score = 1.0
            else:
                # Compute category counts for each row
                category_counts = np.apply_along_axis(lambda x: np.bincount(x, minlength=2), axis=1, arr=ratings)

                # Compute Fleiss' Kappa
                kappa_score = fleiss_kappa(category_counts, method='fleiss')

                # Handle potential NaN values (due to zero variance)
                kappa_score = np.nan_to_num(kappa_score, nan=1.0)  # Assume perfect agreement for uniform cases

            # Store score
            model_kappa_scores[criterion] = round(kappa_score, 3)

        # Append results
        kappa_scores.append(model_kappa_scores)

    # Convert results to DataFrame
    kappa_df = pd.DataFrame(kappa_scores)

    # ✅ **Ensure Numeric Values for Proper Formatting**
    float_columns = ["Clarity", "Grammar", "Relevance", "Bloom’s Level Fit", "Answerability"]
    kappa_df[float_columns] = kappa_df[float_columns].apply(pd.to_numeric)

    # **Export to Excel**
    excel_filename = "Output/Fleiss_Kappa_Scores.xlsx"
    kappa_df.to_excel(excel_filename, index=False)

    st.success(f"Fleiss' Kappa scores exported to `{excel_filename}`.")

    # **Display Model-wise Fleiss' Kappa Scores**
    st.write("### Model-wise Inter-Rater Agreement (Fleiss' Kappa Scores)")
    st.dataframe(kappa_df)  # ✅ Now properly formatted as float

    # **Identify the Best Performing Model**
    model_avg_scores = kappa_df.set_index("Model Name").mean(axis=1)
    best_model = model_avg_scores.idxmax()
    best_score = model_avg_scores.max()

    st.write(f"### 🏆 Best Performing Model: **{best_model}**")
    st.write(f"Average Fleiss' Kappa Score: **{best_score:.3f}**")

    # **Provide Reasoning for Best Model Selection**
    st.write("#### 📌 Why is this the Best Model?")
    st.write(f"- {best_model} achieved the **highest average agreement** across all evaluation criteria.")
    st.write("- Higher Fleiss' Kappa indicates **stronger consensus among SMEs**, meaning the model-generated questions were more consistently evaluated.")
    st.write("- Models with lower agreement might have **more variability in SME judgments**, leading to **uncertain quality assessments**.")

    # **Visualization - Grouped Bar Chart**
    fig, ax = plt.subplots(figsize=(10, 6))
    kappa_df_melted = kappa_df.melt(id_vars=["Model Name"], var_name="Evaluation Criterion", value_name="Fleiss' Kappa Score")
    sns.barplot(data=kappa_df_melted, x="Evaluation Criterion", y="Fleiss' Kappa Score", hue="Model Name", ax=ax)

    # Reference lines for interpretation
    ax.axhline(y=0.61, color="green", linestyle="--", label="Substantial Agreement (0.61)")
    ax.axhline(y=0.80, color="blue", linestyle="--", label="Almost Perfect Agreement (0.80)")

    ax.set_title("Fleiss' Kappa for SME Agreement Across Models")
    ax.set_ylabel("Kappa Score")
    ax.set_xticklabels(ax.get_xticklabels(), rotation=15)
    ax.legend(loc='lower center', ncol=3, bbox_to_anchor=(0.5, -0.3))

    # **Display the plot in Streamlit**
    st.pyplot(fig)

    return kappa_df


# Function to compute majority vote for SME ratings
def compute_final_sme_scores(sme_df):
    """Computes the final SME evaluation scores by majority vote or averaging."""
    
    sme_final_scores = pd.DataFrame()
    sme_final_scores["Sr.No."] = sme_df["Sr.No."]
    sme_final_scores["Model Name"] = sme_df["Model Name"]
    sme_final_scores["Question"] = sme_df["Question"]

    # Compute majority vote (or average if needed)
    for metric in ["Clarity", "Grammar", "Relevance", "Bloom's Level Fit", "Answerable"]:
        sme_columns = [f"{metric}_SME1", f"{metric}_SME2", f"{metric}_SME3"]
        sme_final_scores[metric] = sme_df[sme_columns].mode(axis=1)[0]  # Majority vote
    
    return sme_final_scores



def compare_sme_vs_automated(sme_df, automated_df):
    """
    Compares SME evaluation with automated evaluation for each model.
    Computes Pearson correlation, generates a results table, and identifies the best-performing model.
    """

    st.write("### 📊 Comparing SME vs Automated Evaluations")

    # Ensure column names are correctly aligned before merging
    merge_cols = ["Model Name", "Chapter Name", "Bloom's Level", "Question"]

    # Merge datasets on common keys
    merged_df = pd.merge(sme_df, automated_df, on=merge_cols, how="outer")

    # Compute SME average scores across 3 evaluators
    merged_df["SME Clarity Avg"] = merged_df[["Clarity_SME1", "Clarity_SME2", "Clarity_SME3"]].mean(axis=1)
    merged_df["SME Grammar Avg"] = merged_df[["Grammar_SME1", "Grammar_SME2", "Grammar_SME3"]].mean(axis=1)
    merged_df["SME Relevance Avg"] = merged_df[["Relevance_SME1", "Relevance_SME2", "Relevance_SME3"]].mean(axis=1)
    merged_df["SME Bloom Fit Avg"] = merged_df[["Bloom's Level Fit_SME1", "Bloom's Level Fit_SME2", "Bloom's Level Fit_SME3"]].mean(axis=1)
    merged_df["SME Answerability Avg"] = merged_df[["Answerable_SME1", "Answerable_SME2", "Answerable_SME3"]].mean(axis=1)

    # Define the metric mappings
    metric_pairs = {
        "SME Clarity Avg": "Clarity",
        "SME Grammar Avg": "Grammar",
        "SME Relevance Avg": "Relevance",
        "SME Bloom Fit Avg": "Bloom's Level Fit",
        "SME Answerability Avg": "Answerable"
    }

    # Ensure metric pairs exist in automated_df before using
    metric_pairs = {k: v for k, v in metric_pairs.items() if v in automated_df.columns}

    # Get unique models from dataset
    models = merged_df["Model Name"].unique()

    # Dictionary to store SME parameter-wise correlation results
    sme_param_results = {sme_metric: {} for sme_metric in metric_pairs.keys()}

    for model in models:
        model_df = merged_df[merged_df["Model Name"] == model]

        for sme_avg_col, automated_col in metric_pairs.items():
            valid_data = model_df[[sme_avg_col, automated_col]].dropna()

            if len(valid_data) > 1:  # Ensure enough data points
                pearson_corr, _ = pearsonr(valid_data[sme_avg_col], valid_data[automated_col])
                sme_param_results[sme_avg_col][model] = round(max(0, pearson_corr), 3)  # Ensure non-negative values
            else:
                sme_param_results[sme_avg_col][model] = np.nan  # Not enough data

    # Convert results to a DataFrame
    sme_correlation_df = pd.DataFrame(sme_param_results)

    # ✅ Ensure numeric values for proper formatting
    sme_correlation_df = sme_correlation_df.apply(pd.to_numeric)

    # **Export to Excel**
    excel_filename = "Output/SME_vs_Automated_Correlation.xlsx"
    sme_correlation_df.to_excel(excel_filename, index=True)

    st.success(f"Pearson correlation scores exported to `{excel_filename}`.")

    # **Display Model-wise SME vs Automated Evaluation Correlation Table**
    st.write("### Model-wise SME vs Automated Evaluation Correlation Table")
    st.dataframe(sme_correlation_df.style.format("{:.3f}"))

    # **Identify the Best Performing Model**
    model_avg_scores = sme_correlation_df.mean(axis=1)
    best_model = model_avg_scores.idxmax()
    best_score = model_avg_scores.max()

    st.write(f"### 🏆 Best Performing Model: **{best_model}**")
    st.write(f"Average Pearson Correlation Score: **{best_score:.3f}**")

    # **Provide Reasoning for Best Model Selection**
    st.write("#### 📌 Why is this the Best Model?")
    st.write(f"- {best_model} achieved the **highest correlation** between SME and Automated evaluations.")
    st.write("- A higher Pearson correlation suggests that **automated metrics closely align with SME evaluations**.")
    st.write("- A lower correlation might indicate **discrepancies between human judgment and automated evaluation methods**.")

    # **Visualization - Grouped Bar Chart**
    st.write("### Model-wise SME vs Automated Evaluation Comparison")

    fig, ax = plt.subplots(figsize=(10, 4))

    x_labels = list(sme_correlation_df.index)  # Model Names
    x = np.arange(len(x_labels))  # X positions

    width = 0.15  # Width of bars

    # Plot each SME parameter
    for i, metric in enumerate(metric_pairs.keys()):
        values = sme_correlation_df[metric].values
        ax.bar(x + (i * width), values, width=width, label=metric)

    # Set labels and title
    ax.set_xlabel('Model Name')
    ax.set_ylabel('Pearson Correlation Coefficient')
    ax.set_title('Model-wise SME vs Automated Evaluation Comparison')
    ax.set_xticks(x + width)
    ax.set_xticklabels(x_labels)  # Aligning labels in the center
    #ax.set_ylim(0, 1)  # Ensure non-negative values only
    ax.legend(title="SME Metrics", bbox_to_anchor=(1.05, 1), loc='upper left')

    # Adding the data labels on top of the bars
    for i, metric in enumerate(metric_pairs.keys()):
        for j, value in enumerate(sme_correlation_df[metric].values):
            ax.text(x[j] + (i * width), value, f"{value:.2f}", ha='center', va='bottom', fontsize=8)

    plt.tight_layout() 

    # **Display the plot in Streamlit**
    st.pyplot(fig)

    return sme_correlation_df

import krippendorff
import rbo

def compute_krippendorff_alpha(sme_df):
    """
    Computes Krippendorff’s Alpha for each evaluation criterion per model.
    
    Args:
        sme_df (pd.DataFrame): DataFrame containing SME evaluations.
    
    Returns:
        pd.DataFrame: Krippendorff’s Alpha scores per model.
    """
    evaluation_criteria = {
        "Clarity": ["Clarity_SME1", "Clarity_SME2", "Clarity_SME3"],
        "Grammar": ["Grammar_SME1", "Grammar_SME2", "Grammar_SME3"],
        "Relevance": ["Relevance_SME1", "Relevance_SME2", "Relevance_SME3"],
        "Bloom’s Level Fit": ["Bloom's Level Fit_SME1", "Bloom's Level Fit_SME2", "Bloom's Level Fit_SME3"],
        "Answerability": ["Answerable_SME1", "Answerable_SME2", "Answerable_SME3"]
    }

    alpha_scores = []

    for model in sme_df["Model Name"].unique():
        model_data = sme_df[sme_df["Model Name"] == model]
        model_alpha_scores = {"Model Name": model}

        for criterion, sme_columns in evaluation_criteria.items():
            ratings = model_data[sme_columns].values.T  # Transpose for Krippendorff's Alpha

            # **Check for uniform ratings (only one unique value)**
            unique_values = np.unique(ratings)
            if len(unique_values) <= 1:  # Not enough variance for Krippendorff’s Alpha
                alpha = 1.0  # Assume perfect agreement since all values are the same
            else:
                alpha = krippendorff.alpha(reliability_data=ratings, level_of_measurement="nominal")

            model_alpha_scores[criterion] = round(alpha, 3)

        alpha_scores.append(model_alpha_scores)

    alpha_df = pd.DataFrame(alpha_scores)

    # ✅ Ensure numeric values for formatting
    float_columns = ["Clarity", "Grammar", "Relevance", "Bloom’s Level Fit", "Answerability"]
    alpha_df[float_columns] = alpha_df[float_columns].apply(pd.to_numeric)

    # **Export Krippendorff’s Alpha scores**
    excel_filename = "Output/Krippendorff_Alpha_Scores.xlsx"
    alpha_df.to_excel(excel_filename, index=False)
    
    st.success("Krippendorff’s Alpha scores exported to `Krippendorff_Alpha_Scores.xlsx`.")
    
    # **Display results**
    st.write("### Model-wise Krippendorff’s Alpha Scores")
    st.dataframe(alpha_df)

    # **Identify Best Model**
    model_avg_scores = alpha_df.set_index("Model Name").mean(axis=1)
    best_model = model_avg_scores.idxmax()
    best_score = model_avg_scores.max()
    
    st.write(f"### 🏆 Best Model by Krippendorff’s Alpha: **{best_model}**")
    st.write(f"Average Krippendorff’s Alpha Score: **{best_score:.3f}**")

    return alpha_df


# Streamlit Workflow
st.title("Question Generation and Evaluation")
url = st.text_input("Enter the chapter URL from OpenStax.org")

if st.button("Generate Questions"):
    if url:
        combined_df = generate_questions(url)
        st.success("Question generation completed. Data saved to Excel.")

if st.button("Evaluate Models"):
    #if "vectors" in st.session_state:
    #    combined_df = pd.read_excel(file_path)
    #    evaluate_models(combined_df)
    #else:
    #    st.error("Please generate questions first.") 
    
    if os.path.exists("Output/generated_questions_SME_Response.xlsx"):
        sme_df = pd.read_excel("Output/generated_questions_SME_Response.xlsx")
        automated_df = pd.read_excel(file_path)

        evaluate_models(automated_df)

        # Compute SME agreement using Fleiss' Kappa
        kappa_results = compute_fleiss_kappa_per_model(sme_df)
        
        # Compare SME and Automated Evaluation
        correlation_results = compare_sme_vs_automated(sme_df, automated_df)

        # Compute Krippendorff’s Alpha
        alpha_results = compute_krippendorff_alpha(sme_df)


    else:
        st.error("SME response file not found.")